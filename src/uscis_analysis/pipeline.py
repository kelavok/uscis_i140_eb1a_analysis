from __future__ import annotations

import argparse
import csv
import hashlib
import html.parser
import json
import os
import re
import sys
import urllib.parse
import urllib.request
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

PARSER_VERSION = "i140-mvp-2026-05-22"

KNOWN_ARCHIVE_URLS = [
    "https://www.uscis.gov/sites/default/files/document/data/i140_fy24_q1.xlsx",
    "https://www.uscis.gov/sites/default/files/document/reports/i140_fy24_q2.xlsx",
    "https://www.uscis.gov/sites/default/files/document/data/i140_fy2024_q3.xlsx",
    "https://www.uscis.gov/sites/default/files/document/data/i140_fy24_q3.xlsx",
    "https://www.uscis.gov/sites/default/files/document/data/i140_fy2024_q4.xlsx",
    "https://www.uscis.gov/sites/default/files/document/data/i140_fy2025_q1.xlsx",
    "https://www.uscis.gov/sites/default/files/document/data/i140_fy2025_q2.xlsx",
    "https://www.uscis.gov/sites/default/files/document/data/i140_fy2025_q3_0.xlsx",
    "https://www.uscis.gov/sites/default/files/document/data/i140_fy2025_q4_v1.xlsx",
    "https://www.uscis.gov/sites/default/files/document/data/i140_rec_by_class_country_fy2024_q1.xlsx",
    "https://www.uscis.gov/sites/default/files/document/reports/i140_rec_by_class_country_fy2024_q2.xlsx",
    "https://www.uscis.gov/sites/default/files/document/data/i140_rec_by_class_country_fy2024_q3.xlsx",
    "https://www.uscis.gov/sites/default/files/document/data/i140_rec_by_class_country_fy2024_q4.xlsx",
    "https://www.uscis.gov/sites/default/files/document/data/i140_rec_by_class_country_fy2025_q1.xlsx",
    "https://www.uscis.gov/sites/default/files/document/data/i140_rec_by_class_country_fy2025_q2.xlsx",
    "https://www.uscis.gov/sites/default/files/document/data/i140_rec_by_class_country_fy2025_q3.xlsx",
    "https://www.uscis.gov/sites/default/files/document/data/i140_rec_by_class_country_fy2025_q4_v1.xlsx",
    "https://www.uscis.gov/sites/default/files/document/data/I140_FY22_Q4_RADP.csv",
    "https://www.uscis.gov/sites/default/files/document/data/I140_rec_by_class_country_FY2022_Q3_Q4.csv",
    "https://www.uscis.gov/sites/default/files/document/data/I140_FY22_Q3_RADP_Summary.csv",
    "https://www.uscis.gov/sites/default/files/document/data/I140_rec_by_class_country_FY2022_Q1_Q2.csv",
    "https://www.uscis.gov/sites/default/files/document/reports/I140_rec_by_class_country_FY2021_Q3_Q4.csv",
    "https://www.uscis.gov/sites/default/files/document/data/I140_FY23_Q1_RADP.csv",
    "https://www.uscis.gov/sites/default/files/document/data/I140_rec_by_class_country_FY2023_Q1.csv",
    "https://www.uscis.gov/sites/default/files/document/data/I140_rec_by_class_country_FY2023_Q1_Q2.csv",
    "https://www.uscis.gov/sites/default/files/document/data/i140_fy23_q4_radp.csv",
    "https://www.uscis.gov/sites/default/files/document/data/i140_rec_by_class_country_fy2023_q4.csv",
]


@dataclass
class SourceLink:
    title: str
    url: str
    file_name: str
    report_family: str
    report_fiscal_year: int | None
    report_quarter: int | None
    fiscal_year_label: str | None


class USCISLinkParser(html.parser.HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.links: list[tuple[str, str]] = []
        self._href: str | None = None
        self._text: list[str] = []

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        if tag.lower() != "a":
            return
        attr = dict(attrs)
        href = attr.get("href")
        if href:
            self._href = href
            self._text = []

    def handle_data(self, data: str) -> None:
        if self._href:
            self._text.append(data)

    def handle_endtag(self, tag: str) -> None:
        if tag.lower() == "a" and self._href:
            text = normalize_space(" ".join(self._text))
            self.links.append((self._href, text))
            self._href = None
            self._text = []


def normalize_space(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).replace("\xa0", " ")).strip()


def load_env(root: Path) -> dict[str, str]:
    env_path = root / ".env"
    values = dict(os.environ)
    if env_path.exists():
        for line in env_path.read_text(encoding="utf-8").splitlines():
            if not line.strip() or line.lstrip().startswith("#") or "=" not in line:
                continue
            key, value = line.split("=", 1)
            values[key.strip()] = value.strip()
    return values


def fetch_text(url: str) -> str:
    request = urllib.request.Request(url, headers=request_headers())
    with urllib.request.urlopen(request, timeout=60) as response:
        return response.read().decode("utf-8", errors="replace")


def request_headers() -> dict[str, str]:
    return {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        "Accept": "text/html,application/xhtml+xml,application/xml,application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,*/*",
        "Referer": "https://www.uscis.gov/tools/reports-and-studies/immigration-and-citizenship-data",
    }


def discover_sources(index_url: str) -> list[SourceLink]:
    html = fetch_text(index_url)
    parser = USCISLinkParser()
    parser.feed(html)
    sources: list[SourceLink] = []
    seen: set[str] = set()

    for href, title in parser.links:
        if ".xlsx" not in href.lower() or "i-140" not in title.lower():
            continue
        family = detect_report_family(title)
        if not family:
            continue
        absolute_url = urllib.parse.urljoin(index_url, href)
        if absolute_url in seen:
            continue
        seen.add(absolute_url)
        file_name = Path(urllib.parse.urlparse(absolute_url).path).name
        report_fy, report_q, fy_label = parse_report_period(title)
        sources.append(
            SourceLink(
                title=title,
                url=absolute_url,
                file_name=file_name,
                report_family=family,
                report_fiscal_year=report_fy,
                report_quarter=report_q,
                fiscal_year_label=fy_label,
            )
        )
    return sources


def discover_sources_with_probes(index_url: str, start_year: int, end_year: int) -> list[SourceLink]:
    sources: list[SourceLink] = []
    index_urls = re.split(r"[\n;]+", index_url)
    for url in [part.strip() for part in index_urls if part.strip()]:
        try:
            sources.extend(discover_sources(url))
        except Exception as exc:
            print(f"USCIS index discovery failed for {url} ({exc}); continuing")
    for known in known_archive_sources():
        if known.url not in {source.url for source in sources}:
            sources.append(known)
    seen_urls = {source.url for source in sources}
    candidates = candidate_archive_sources(start_year, end_year)

    def exists(source: SourceLink) -> SourceLink | None:
        request = urllib.request.Request(source.url, method="HEAD", headers=request_headers())
        try:
            with urllib.request.urlopen(request, timeout=10) as response:
                if response.status == 200:
                    return source
        except Exception:
            return None
        return None

    with ThreadPoolExecutor(max_workers=16) as executor:
        futures = [executor.submit(exists, candidate) for candidate in candidates if candidate.url not in seen_urls]
        for future in as_completed(futures):
            source = future.result()
            if not source or source.url in seen_urls:
                continue
            seen_urls.add(source.url)
            sources.append(source)

    return sorted(
        sources,
        key=lambda s: (
            s.report_fiscal_year or 0,
            s.report_quarter or 0,
            s.report_family,
            s.file_name,
        ),
    )


def discover_local_hand_sources(root: Path) -> list[SourceLink]:
    raw_dir = root / "data" / "raw"
    hand_dir = raw_dir / "raw_by_hands"
    files = []
    if hand_dir.exists():
        files.extend(path for path in sorted(hand_dir.glob("*")) if path.is_file())
    if raw_dir.exists():
        files.extend(path for path in sorted(raw_dir.glob("*")) if path.is_file())
    if not files:
        return []
    preferred_stems = {
        normalize_file_stem(path.name)
        for path in files
        if path.suffix.lower() in {".csv", ".xlsx"}
    }
    seen_names: set[str] = set()
    sources: list[SourceLink] = []
    for path in files:
        if path.suffix.lower() not in {".pdf", ".csv", ".xlsx"}:
            continue
        if path.suffix.lower() == ".pdf" and normalize_file_stem(path.name) in preferred_stems:
            continue
        name_key = normalize_file_stem(path.name)
        if name_key in seen_names:
            continue
        seen_names.add(name_key)
        parsed = parse_file_name_period(path.name)
        if not parsed:
            continue
        family, year, quarter = parsed
        title = path.stem.replace("_", " ")
        local_scheme = "local_hand" if path.parent == hand_dir else "local_raw"
        sources.append(
            SourceLink(
                title=title,
                url=f"{local_scheme}://{path.name}",
                file_name=path.name,
                report_family=family,
                report_fiscal_year=year,
                report_quarter=quarter,
                fiscal_year_label=f"Fiscal Year {year}, Quarter {quarter}",
            )
        )
    return sources


def normalize_file_stem(name: str) -> str:
    stem = Path(name).stem.lower()
    stem = re.sub(r"\s*\(\d+\)$", "", stem)
    stem = stem.replace("-", "_")
    return re.sub(r"[^a-z0-9]+", "_", stem).strip("_")


def candidate_archive_sources(start_year: int, end_year: int) -> list[SourceLink]:
    bases = [
        "https://www.uscis.gov/sites/default/files/document/data/",
        "https://www.uscis.gov/sites/default/files/document/reports/",
    ]
    candidates: list[SourceLink] = []
    for year in range(start_year, end_year + 1):
        short_year = f"{year % 100:02d}"
        for quarter in range(1, 5):
            for family, title_template, names in [
                (
                    "fy_quarter_status",
                    "Form I-140 by Fiscal Year, Quarter and Case Status (Fiscal Year {year}, Quarter {quarter})",
                    [
                        f"i140_fy{year}_q{quarter}.xlsx",
                        f"i140_fy{year}_q{quarter}_v1.xlsx",
                        f"i140_fy{short_year}_q{quarter}.xlsx",
                        f"i140_fy{short_year}_q{quarter}_v1.xlsx",
                    ],
                ),
                (
                    "preference_country",
                    "Form I-140, Receipts and Current Status by Preference and Country (Fiscal Year {year}, Quarter {quarter})",
                    [
                        f"i140_rec_by_class_country_fy{year}_q{quarter}.xlsx",
                        f"i140_rec_by_class_country_fy{year}_q{quarter}_v1.xlsx",
                        f"i140_rec_by_class_country_fy{short_year}_q{quarter}.xlsx",
                        f"i140_rec_by_class_country_fy{short_year}_q{quarter}_v1.xlsx",
                    ],
                ),
            ]:
                title = title_template.format(year=year, quarter=quarter)
                for base in bases:
                    for file_name in names:
                        candidates.append(
                            SourceLink(
                                title=title,
                                url=base + file_name,
                                file_name=file_name,
                                report_family=family,
                                report_fiscal_year=year,
                                report_quarter=quarter,
                                fiscal_year_label=f"Fiscal Year {year}, Quarter {quarter}",
                            )
                        )
    return candidates


def known_archive_sources() -> list[SourceLink]:
    sources: list[SourceLink] = []
    for url in KNOWN_ARCHIVE_URLS:
        file_name = Path(urllib.parse.urlparse(url).path).name
        parsed = parse_file_name_period(file_name)
        if not parsed:
            continue
        family, year, quarter = parsed
        if family == "fy_quarter_status":
            title = f"Form I-140 by Fiscal Year, Quarter and Case Status (Fiscal Year {year}, Quarter {quarter})"
        else:
            title = f"Form I-140, Receipts and Current Status by Preference and Country (Fiscal Year {year}, Quarter {quarter})"
        sources.append(
            SourceLink(
                title=title,
                url=url,
                file_name=file_name,
                report_family=family,
                report_fiscal_year=year,
                report_quarter=quarter,
                fiscal_year_label=f"Fiscal Year {year}, Quarter {quarter}",
            )
        )
    return sources


def parse_file_name_period(file_name: str) -> tuple[str, int, int] | None:
    lowered = file_name.lower()
    if "quarterly_all_forms" in lowered or "ecn_1893" in lowered:
        family = "all_forms"
    elif re.search(r"(^|[_-])(rec|app)_(cob|country)", lowered) or re.search(r"fy\d{2}_q\d_(rec|app)_cob", lowered):
        family = "country_category"
    elif "rec_by_class_country" in lowered or "by_class_country" in lowered:
        family = "preference_country"
    elif re.match(r"i-?140_fy", lowered) and ("radp" in lowered or lowered.endswith(".xlsx") or re.search(r"_q[1-4]\.csv$", lowered)):
        family = "fy_quarter_status"
    else:
        return None
    if "fy09_19" in lowered:
        return family, 2019, 4
    all_forms_match = re.search(r"fy(20\d{2}|\d{2})_?q([1-4])|fy(20\d{2}|\d{2})q([1-4])", lowered)
    if all_forms_match and family == "all_forms":
        year_text = all_forms_match.group(1) or all_forms_match.group(3)
        quarter_text = all_forms_match.group(2) or all_forms_match.group(4)
        year = int(year_text) if len(year_text) == 4 else 2000 + int(year_text)
        return family, year, int(quarter_text)
    match = re.search(r"fy(20\d{2}|\d{2})_q([1-4])", lowered)
    range_match = re.search(r"fy(20\d{2}|\d{2})_q([1-4])_q([1-4])", lowered)
    if not match and not range_match:
        return None
    match = range_match or match
    year_text = match.group(1)
    year = int(year_text) if len(year_text) == 4 else 2000 + int(year_text)
    quarter = int(match.group(3) if range_match else match.group(2))
    return family, year, quarter


def fiscal_quarter_end(fiscal_year: int | None, quarter: int | None) -> date:
    if not fiscal_year or not quarter:
        return date.today()
    if quarter == 1:
        return date(fiscal_year - 1, 12, 31)
    if quarter == 2:
        return date(fiscal_year, 3, 31)
    if quarter == 3:
        return date(fiscal_year, 6, 30)
    return date(fiscal_year, 9, 30)


def detect_report_family(title: str) -> str | None:
    lowered = title.lower()
    if "form i-140 by fiscal year, quarter and case status" in lowered:
        return "fy_quarter_status"
    if "form i-140, receipts and current status by preference and country" in lowered:
        return "preference_country"
    return None


def parse_report_period(title: str) -> tuple[int | None, int | None, str | None]:
    match = re.search(r"Fiscal Year\s+(\d{4})(?:,\s*Quarter\s*(\d))?", title, re.I)
    if not match:
        return None, None, None
    fy = int(match.group(1))
    quarter = int(match.group(2)) if match.group(2) else None
    return fy, quarter, match.group(0)


def download_file(source: SourceLink, raw_dir: Path) -> tuple[Path, str]:
    if source.url.startswith("local_hand://"):
        path = raw_dir / "raw_by_hands" / source.file_name
        data = path.read_bytes()
        return path, hashlib.sha256(data).hexdigest()
    if source.url.startswith("local_raw://"):
        path = raw_dir / source.file_name
        data = path.read_bytes()
        return path, hashlib.sha256(data).hexdigest()
    raw_dir.mkdir(parents=True, exist_ok=True)
    path = raw_dir / source.file_name
    if path.exists():
        data = path.read_bytes()
        return path, hashlib.sha256(data).hexdigest()
    request = urllib.request.Request(source.url, headers=request_headers())
    try:
        with urllib.request.urlopen(request, timeout=120) as response:
            data = response.read()
    except Exception:
        data = download_from_wayback(source.url)
    path.write_bytes(data)
    return path, hashlib.sha256(data).hexdigest()


def download_from_wayback(url: str) -> bytes:
    cdx_url = (
        "https://web.archive.org/cdx?"
        + urllib.parse.urlencode(
            {
                "url": urllib.parse.urlparse(url).netloc + urllib.parse.urlparse(url).path,
                "output": "json",
                "fl": "timestamp,original,statuscode,mimetype,digest",
                "filter": "statuscode:200",
                "collapse": "digest",
            }
        )
    )
    request = urllib.request.Request(cdx_url, headers=request_headers())
    with urllib.request.urlopen(request, timeout=60) as response:
        rows = json.loads(response.read().decode("utf-8"))
    if len(rows) < 2:
        raise RuntimeError(f"No Wayback snapshot found for {url}")
    timestamp = rows[-1][0]
    archived_url = f"https://web.archive.org/web/{timestamp}id_/{url}"
    request = urllib.request.Request(archived_url, headers=request_headers())
    with urllib.request.urlopen(request, timeout=120) as response:
        data = response.read()
    if url.lower().endswith(".xlsx") and not data.startswith(b"PK"):
        raise RuntimeError(f"Wayback snapshot is not an XLSX file for {url}")
    return data


def merged_lookup(ws) -> dict[str, tuple[Any, str]]:
    lookup: dict[str, tuple[Any, str]] = {}
    for merged in ws.merged_cells.ranges:
        top_left = ws.cell(merged.min_row, merged.min_col).value
        for row in range(merged.min_row, merged.max_row + 1):
            for col in range(merged.min_col, merged.max_col + 1):
                lookup[f"{get_column_letter(col)}{row}"] = (top_left, str(merged))
    return lookup


def cell_value(ws, row: int, col: int, merged: dict[str, tuple[Any, str]]) -> Any:
    address = f"{get_column_letter(col)}{row}"
    if address in merged:
        return merged[address][0]
    return ws.cell(row, col).value


def clean_count(value: Any) -> int | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return int(value)
    text = normalize_space(value)
    if not text:
        return None
    if text in {"-", "—", "–"}:
        return 0
    text = re.sub(r"[^\d.-]", "", text)
    if not text:
        return None
    return int(float(text))


def status_code(label: str) -> str | None:
    text = normalize_space(label).lower()
    text = re.sub(r"\d+", "", text).strip()
    if text.startswith("received") or text.startswith("forms received"):
        return "received"
    if text.startswith("approved"):
        return "approved"
    if text.startswith("denied"):
        return "denied"
    if text.startswith("pending, other"):
        return "pending_other"
    if text.startswith("pending"):
        return "pending"
    if text.startswith("total petition"):
        return "received"
    return None


def category_code(label: str) -> str | None:
    text = normalize_space(label)
    lowered = text.lower()
    if lowered == "total":
        return "TOTAL"
    if "first preference" in lowered:
        return "EB1"
    if "second preference" in lowered:
        return "EB2"
    if "third preference" in lowered:
        return "EB3"
    if "other and unknown" in lowered:
        return "OTHER_UNKNOWN"
    paren = re.findall(r"\(([A-Z0-9]{2,4})\)", text)
    if paren:
        return paren[-1].upper()
    if "national interest waiver" in lowered:
        return "NIW"
    return None


def quarter_number(label: str) -> int | None:
    lowered = normalize_space(label).lower()
    if "1st" in lowered or "first" in lowered:
        return 1
    if "2nd" in lowered or "second" in lowered:
        return 2
    if "3rd" in lowered or "third" in lowered:
        return 3
    if "4th" in lowered or "fourth" in lowered:
        return 4
    return None


def country_from_sheet(sheet_name: str) -> tuple[str, str]:
    normalized = normalize_space(sheet_name)
    normalized = re.sub(r"\b(?:FY)?(?:20)?\d{2}\b", "", normalized, flags=re.I)
    normalized = normalize_space(normalized)
    if normalized.lower() in {"all countries", "base status", "app status", "readp summary"}:
        return "ALL", "All Countries"
    code = re.sub(r"[^A-Za-z0-9]+", "_", normalized).strip("_").upper()[:40]
    return code or "ALL", normalized or "All Countries"


def parse_workbook(source: SourceLink, path: Path, file_hash: str, snapshot_date: date) -> tuple[list[dict], list[dict], list[dict], list[dict], list[dict], list[dict]]:
    wb = load_workbook(path, data_only=True)
    source_rows = [{
        "file_hash_sha256": file_hash,
        "source_name": "USCIS",
        "form_type": "ALL_FORMS" if source.report_family == "all_forms" else "I-140",
        "report_family": source.report_family,
        "report_title": source.title,
        "file_name": source.file_name,
        "source_url": source.url,
        "local_path": str(path),
        "report_fiscal_year": source.report_fiscal_year or "",
        "report_quarter": source.report_quarter or "",
        "fiscal_year_label": source.fiscal_year_label or "",
        "published_date": "",
        "snapshot_date": snapshot_date.isoformat(),
        "parser_version": PARSER_VERSION,
        "parse_status": "parsed",
        "parse_notes": "",
    }]
    sheets: list[dict] = []
    cells: list[dict] = []
    facts: list[dict] = []
    form_facts: list[dict] = []
    countries: list[dict] = []

    for idx, ws in enumerate(wb.worksheets):
        merged = merged_lookup(ws)
        detected = source.report_family
        sheets.append({
            "file_hash_sha256": file_hash,
            "sheet_index": idx,
            "sheet_name": ws.title,
            "max_row": ws.max_row,
            "max_column": ws.max_column,
            "detected_table_type": detected,
            "detection_confidence": "0.9500",
        })
        for row in ws.iter_rows():
            for cell in row:
                raw = cell.value
                address = cell.coordinate
                is_merged = address in merged
                merged_range = merged[address][1] if is_merged else ""
                effective = merged[address][0] if is_merged else raw
                if effective is None:
                    continue
                cells.append({
                    "file_hash_sha256": file_hash,
                    "sheet_index": idx,
                    "row_num": cell.row,
                    "col_num": cell.column,
                    "cell_address": address,
                    "raw_value": "" if raw is None else str(raw),
                    "normalized_value": normalize_space(effective),
                    "is_merged": "true" if is_merged else "false",
                    "merged_range": merged_range,
                })
        if source.report_family == "fy_quarter_status":
            facts.extend(parse_fy_quarter_sheet(source, ws, merged, file_hash, idx, snapshot_date))
        elif source.report_family == "preference_country":
            country_code_value, country_name = country_from_sheet(ws.title)
            countries.append({"country_code": country_code_value, "country_name": country_name, "uscis_country_label": country_name})
            facts.extend(parse_preference_country_sheet(source, ws, merged, file_hash, idx, snapshot_date, country_code_value))
        elif source.report_family == "all_forms":
            grid = [[cell_value(ws, row, col, merged) for col in range(1, ws.max_column + 1)] for row in range(1, ws.max_row + 1)]
            form_facts.extend(parse_all_forms_grid(source, grid, file_hash, idx, snapshot_date))
        elif source.report_family == "country_category":
            grid = [[ws.cell(row, col).value for col in range(1, ws.max_column + 1)] for row in range(1, ws.max_row + 1)]
            facts.extend(parse_country_category_grid(source, grid, file_hash, idx, snapshot_date))
    return source_rows, sheets, cells, facts, countries, form_facts


def parse_source_file(source: SourceLink, path: Path, file_hash: str, snapshot_date: date) -> tuple[list[dict], list[dict], list[dict], list[dict], list[dict], list[dict]]:
    if path.suffix.lower() == ".pdf":
        return parse_pdf_source(source, path, file_hash, snapshot_date)
    if path.suffix.lower() == ".csv":
        return parse_csv_source(source, path, file_hash, snapshot_date)
    return parse_workbook(source, path, file_hash, snapshot_date)


def parse_pdf_source(source: SourceLink, path: Path, file_hash: str, snapshot_date: date) -> tuple[list[dict], list[dict], list[dict], list[dict], list[dict], list[dict]]:
    from pypdf import PdfReader

    reader = PdfReader(str(path))
    page_texts = [page.extract_text() or "" for page in reader.pages]
    lines: list[str] = []
    cells: list[dict] = []
    for page_idx, text in enumerate(page_texts, start=1):
        for line_idx, line in enumerate(text.splitlines(), start=1):
            normalized = normalize_space(line)
            if not normalized:
                continue
            lines.append(normalized)
            cells.append({
                "file_hash_sha256": file_hash,
                "sheet_index": 0,
                "row_num": len(lines),
                "col_num": 1,
                "cell_address": f"P{page_idx}:L{line_idx}",
                "raw_value": line,
                "normalized_value": normalized,
                "is_merged": "false",
                "merged_range": "",
            })
    source_rows = [{
        "file_hash_sha256": file_hash,
        "source_name": "USCIS",
        "form_type": "ALL_FORMS" if source.report_family == "all_forms" else "I-140",
        "report_family": source.report_family,
        "report_title": source.title,
        "file_name": source.file_name,
        "source_url": source.url,
        "local_path": str(path),
        "report_fiscal_year": source.report_fiscal_year or "",
        "report_quarter": source.report_quarter or "",
        "fiscal_year_label": source.fiscal_year_label or "",
        "published_date": "",
        "snapshot_date": snapshot_date.isoformat(),
        "parser_version": PARSER_VERSION,
        "parse_status": "parsed",
        "parse_notes": "pdf_text_source",
    }]
    sheets = [{
        "file_hash_sha256": file_hash,
        "sheet_index": 0,
        "sheet_name": "PDF Text",
        "max_row": len(lines),
        "max_column": 1,
        "detected_table_type": source.report_family,
        "detection_confidence": "0.7000",
    }]
    form_facts: list[dict] = []
    if source.report_family == "preference_country":
        facts, countries = parse_pdf_preference_country(source, lines, file_hash, snapshot_date)
    elif source.report_family == "all_forms":
        facts = []
        countries = [{"country_code": "ALL", "country_name": "All Countries", "uscis_country_label": "All Countries"}]
        form_facts = parse_pdf_all_forms(source, lines, file_hash, snapshot_date)
    else:
        facts, countries = [], [{"country_code": "ALL", "country_name": "All Countries", "uscis_country_label": "All Countries"}]
    return source_rows, sheets, cells, facts, countries, form_facts


def parse_csv_source(source: SourceLink, path: Path, file_hash: str, snapshot_date: date) -> tuple[list[dict], list[dict], list[dict], list[dict], list[dict], list[dict]]:
    rows = list(csv.reader(path.read_text(encoding="utf-8-sig", errors="replace").splitlines()))
    source_rows = [{
        "file_hash_sha256": file_hash,
        "source_name": "USCIS",
        "form_type": "ALL_FORMS" if source.report_family == "all_forms" else "I-140",
        "report_family": source.report_family,
        "report_title": source.title,
        "file_name": source.file_name,
        "source_url": source.url,
        "local_path": str(path),
        "report_fiscal_year": source.report_fiscal_year or "",
        "report_quarter": source.report_quarter or "",
        "fiscal_year_label": source.fiscal_year_label or "",
        "published_date": "",
        "snapshot_date": snapshot_date.isoformat(),
        "parser_version": PARSER_VERSION,
        "parse_status": "parsed",
        "parse_notes": "csv_source",
    }]
    max_cols = max((len(row) for row in rows), default=0)
    sheets = [{
        "file_hash_sha256": file_hash,
        "sheet_index": 0,
        "sheet_name": "CSV",
        "max_row": len(rows),
        "max_column": max_cols,
        "detected_table_type": source.report_family,
        "detection_confidence": "0.8500",
    }]
    cells = []
    for r_idx, row in enumerate(rows, start=1):
        for c_idx, value in enumerate(row, start=1):
            if normalize_space(value):
                cells.append({
                    "file_hash_sha256": file_hash,
                    "sheet_index": 0,
                    "row_num": r_idx,
                    "col_num": c_idx,
                    "cell_address": f"{get_column_letter(c_idx)}{r_idx}",
                    "raw_value": value,
                    "normalized_value": normalize_space(value),
                    "is_merged": "false",
                    "merged_range": "",
                })
    countries = [{"country_code": "ALL", "country_name": "All Countries", "uscis_country_label": "All Countries"}]
    form_facts: list[dict] = []
    if source.report_family == "fy_quarter_status":
        facts = parse_csv_fy_quarter(source, rows, file_hash, snapshot_date)
    elif source.report_family == "preference_country":
        facts, countries = parse_csv_preference_country(source, rows, file_hash, snapshot_date)
    elif source.report_family == "country_category":
        facts = parse_country_category_grid(source, rows, file_hash, 0, snapshot_date)
    elif source.report_family == "all_forms":
        facts = []
        form_facts = parse_all_forms_grid(source, rows, file_hash, 0, snapshot_date)
    else:
        facts = []
    return source_rows, sheets, cells, facts, countries, form_facts


def csv_cell(rows: list[list[str]], row: int, col: int) -> str:
    if row < 0 or row >= len(rows) or col < 0 or col >= len(rows[row]):
        return ""
    return rows[row][col]


def parse_csv_fy_quarter(source: SourceLink, rows: list[list[str]], file_hash: str, snapshot_date: date) -> list[dict]:
    status_row = next((i for i, row in enumerate(rows) if sum(1 for v in row if status_code(v)) >= 4), None)
    if status_row is None:
        return []
    quarter_row = max(0, status_row - 1)
    facts = []
    for r in range(status_row + 1, len(rows)):
        label = normalize_space(csv_cell(rows, r, 0))
        if not label:
            continue
        if label.lower().startswith(("table key", "references", "source")):
            break
        cat = category_code(label)
        if not cat:
            continue
        for c in range(1, max(len(rows[r]), len(rows[status_row]))):
            stat = status_code(csv_cell(rows, status_row, c))
            if not stat:
                continue
            group_index = ((c - 1) // 4) + 1
            if group_index <= 4:
                q_num = group_index
                q_label = quarter_label(group_index)
                is_total_column = False
            else:
                q_label = "Fiscal Year Total"
                q_num = None
                is_total_column = True
            if q_num and source.report_quarter and q_num > source.report_quarter:
                continue
            count = clean_count(csv_cell(rows, r, c))
            if count is None:
                continue
            facts.append(fact_row(
                file_hash, 0, cat, "ALL", stat, source.report_fiscal_year or 0,
                q_num, source, snapshot_date, count, label,
                f"{q_label} / {csv_cell(rows, status_row, c)}",
                f"{get_column_letter(c + 1)}{r + 1}", cat == "TOTAL" or "preference" in label.lower(), is_total_column
            ))
    return facts


def parse_csv_preference_country(source: SourceLink, rows: list[list[str]], file_hash: str, snapshot_date: date) -> tuple[list[dict], list[dict]]:
    facts = []
    countries: dict[str, dict] = {"ALL": {"country_code": "ALL", "country_name": "All Countries", "uscis_country_label": "All Countries"}}
    section_starts = [i for i, row in enumerate(rows) if any("Current Status for" in str(v) for v in row)]
    if not section_starts:
        section_starts = [0]
    section_starts.append(len(rows))
    for start, end in zip(section_starts, section_starts[1:]):
        section_text = " ".join(rows[start])
        country_name = re.sub(r".*Current Status for\s+", "", section_text, flags=re.I).strip(" ,")
        if not country_name or country_name == section_text:
            country_name = "All Countries"
        country_name = country_name.replace("the Philippines", "Philippines")
        country_code_value, normalized_country = country_from_sheet(country_name)
        countries[country_code_value] = {"country_code": country_code_value, "country_name": normalized_country, "uscis_country_label": normalized_country}
        header_row = None
        for i in range(start, min(end, start + 12)):
            if sum(1 for v in rows[i] if re.fullmatch(r"\s*\d{4}\s*", str(v))) >= 3:
                header_row = i
                break
        if header_row is None:
            continue
        current_category = "TOTAL"
        approval_section = False
        for r in range(header_row + 1, end):
            label = normalize_space(csv_cell(rows, r, 0))
            if not label:
                continue
            lowered = label.lower()
            if lowered.startswith(("table key", "references", "source", "note")):
                break
            if "approvals by category" in lowered:
                approval_section = True
                continue
            cat = category_code(label)
            stat = status_code(label)
            if lowered == "total":
                cat = "TOTAL"
                stat = "received"
            if cat and not stat and not approval_section:
                current_category = cat
                continue
            if cat and approval_section:
                current_category = cat
                stat = "approved"
            if not stat:
                continue
            target_category = cat if cat and stat == "approved" and approval_section else current_category
            for c in range(1, len(rows[header_row])):
                header = normalize_space(csv_cell(rows, header_row, c))
                if not re.fullmatch(r"\d{4}", header):
                    continue
                count = clean_count(csv_cell(rows, r, c))
                if count is None:
                    continue
                facts.append(fact_row(
                    file_hash, 0, target_category, country_code_value, stat, int(header), None,
                    source, snapshot_date, count, label, header,
                    f"{get_column_letter(c + 1)}{r + 1}", target_category == "TOTAL", False
                ))
    return facts, list(countries.values())


def parse_country_category_grid(source: SourceLink, rows: list[list[Any]], file_hash: str, sheet_index: int, snapshot_date: date) -> list[dict]:
    status = "approved" if re.search(r"(^|[_-])app(_|$)|approvals", source.file_name, re.I) else "received"
    header_row = next((i for i, row in enumerate(rows) if any("Beneficiary Country" in normalize_space(v) for v in row)), None)
    if header_row is None:
        return []
    facts: list[dict] = []
    for r in range(header_row + 1, len(rows)):
        country_label = normalize_space(csv_cell_any(rows, r, 0))
        if not country_label or country_label.lower().startswith(("table key", "reference", "note", "source")):
            continue
        country_code_value, _ = country_from_sheet(country_label)
        for c in range(1, len(rows[header_row])):
            raw_header = normalize_space(csv_cell_any(rows, header_row, c))
            cat = category_code(raw_header)
            if not cat:
                cat = category_code_by_position(c)
            if not cat and raw_header.lower() == "total":
                cat = "TOTAL"
            if not cat:
                continue
            count = clean_count(csv_cell_any(rows, r, c))
            if count is None:
                continue
            facts.append(fact_row(
                file_hash, sheet_index, cat, country_code_value, status,
                source.report_fiscal_year or 0, source.report_quarter, source,
                snapshot_date, count, country_label, raw_header,
                f"{get_column_letter(c + 1)}{r + 1}", country_label.lower() == "total", False
            ))
    return facts


def category_code_by_position(col_index: int) -> str | None:
    order = {
        1: "E11",
        2: "E12",
        3: "E13",
        4: "E21",
        5: "NIW",
        6: "E31",
        7: "E32",
        8: "EW3",
        9: "TOTAL",
    }
    return order.get(col_index)


def parse_all_forms_grid(source: SourceLink, rows: list[list[Any]], file_hash: str, sheet_index: int, snapshot_date: date) -> list[dict]:
    header_row = next((i for i, row in enumerate(rows) if any("Category and Form Number" in normalize_space(v) for v in row)), None)
    if header_row is None:
        header_row = next((i for i, row in enumerate(rows) if any("Received" in normalize_space(v) for v in row) and any("Approved" in normalize_space(v) for v in row)), None)
    if header_row is None:
        return []
    header_block_start = max(0, header_row - 3)
    header_block_end = min(len(rows), header_row + 3)
    max_cols = max((len(row) for row in rows[max(0, header_row - 3):min(len(rows), header_row + 3)]), default=0)
    period_labels = all_forms_period_labels(rows, header_block_start, header_row, max_cols)
    if source.report_quarter and (source.report_fiscal_year or 0) >= 2021:
        ytd_start = next((i for i, label in enumerate(period_labels) if "fiscal year" in label.lower()), None)
        if ytd_start is not None:
            q_label = quarter_label(source.report_quarter)
            period_labels = [q_label if i < ytd_start else label for i, label in enumerate(period_labels)]
    status_labels = [
        all_forms_status_for_column(rows, header_block_start, header_block_end, c)
        for c in range(max_cols)
    ]
    facts: list[dict] = []
    current_category = ""
    for r in range(header_row + 1, len(rows)):
        first = normalize_space(csv_cell_any(rows, r, 0))
        second = normalize_space(csv_cell_any(rows, r, 1))
        third = normalize_space(csv_cell_any(rows, r, 2))
        if not first and not second and not third:
            continue
        if first.lower().startswith(("note", "reference", "source", "table key")):
            break
        row_values = [clean_decimal(v) for v in rows[r]]
        has_counts = any(v is not None for v in row_values[2:])
        form_code = normalize_form_code(first or second)
        desc = second if form_code and second else third
        if not form_code and first and not has_counts:
            current_category = first
            continue
        if first.upper().startswith("TOTAL") or first.lower().startswith("total - all forms"):
            form_code = "TOTAL"
            desc = "Total - All Forms"
        if not form_code:
            continue
        for c in range(0, max(max_cols, len(rows[r]))):
            stat = status_labels[c] if c < len(status_labels) else all_forms_status_for_column(rows, header_block_start, header_block_end, c)
            if not stat:
                continue
            value = clean_decimal(csv_cell_any(rows, r, c))
            if value is None:
                continue
            period_label = period_labels[c] if c < len(period_labels) else ""
            q_num = quarter_number(period_label)
            period_scope = "ytd" if "fiscal year" in period_label.lower() or "to-date" in period_label.lower() else "quarter"
            if not q_num and period_scope == "quarter":
                q_num = source.report_quarter
            facts.append({
                "file_hash_sha256": file_hash,
                "sheet_index": sheet_index,
                "form_code": form_code,
                "form_name": desc,
                "status_code": stat,
                "report_fiscal_year": source.report_fiscal_year or "",
                "report_quarter": source.report_quarter or "",
                "snapshot_date": snapshot_date.isoformat(),
                "period_scope": period_scope,
                "period_quarter": q_num or "",
                "count_value": value,
                "form_category": current_category,
                "form_description": desc,
                "raw_row_label": first,
                "raw_column_label": f"{period_label} / {csv_cell_any(rows, header_row, c)}",
                "raw_cell_address": f"{get_column_letter(c + 1)}{r + 1}",
                "extraction_method": "script",
                "extraction_confidence": "0.8500",
                "reviewed": "false",
            })
    return facts


def all_forms_status_for_column(rows: list[list[Any]], start_row: int, end_row: int, col: int) -> str | None:
    text = normalize_space(" ".join(normalize_space(csv_cell_any(rows, r, col)) for r in range(start_row, end_row)))
    lowered = re.sub(r"\d+", "", text).lower()
    if "processing" in lowered or re.search(r"\btime\b", lowered) or re.search(r"\btotal\b", lowered):
        return None
    if "approved" in lowered:
        return "approved"
    if "denied" in lowered:
        return "denied"
    if "pending" in lowered:
        return "pending"
    if "received" in lowered:
        return "received"
    return None


def quarter_label(quarter: int) -> str:
    labels = {
        1: "1st Quarter",
        2: "2nd Quarter",
        3: "3rd Quarter",
        4: "4th Quarter",
    }
    return labels.get(quarter, f"Quarter {quarter}")


def all_forms_period_labels(rows: list[list[Any]], start_row: int, header_row: int, max_cols: int) -> list[str]:
    period_rows = []
    for r in range(start_row, header_row + 1):
        row_text = " ".join(normalize_space(v) for v in rows[r]) if r < len(rows) else ""
        if re.search(r"\b(?:1st|2nd|3rd|4th|first|second|third|fourth)\s+quarter\b", row_text, re.I) or re.search(r"fiscal year\s*-\s*to date|fiscal year\s*-\s*to-date", row_text, re.I):
            period_rows.append(r)
    labels = [""] * max_cols
    if not period_rows:
        return labels

    markers: list[tuple[int, str]] = []
    for r in period_rows:
        for c in range(max_cols):
            label = normalize_space(csv_cell_any(rows, r, c))
            if quarter_number(label) or re.search(r"fiscal year\s*-\s*to date|fiscal year\s*-\s*to-date", label, re.I):
                markers.append((c, label))
    markers = sorted({(c, label) for c, label in markers})
    if not markers:
        return labels

    ytd_markers = [(c, label) for c, label in markers if "fiscal year" in label.lower()]
    quarter_markers = [(c, label) for c, label in markers if quarter_number(label)]
    if len(markers) == 2 and ytd_markers and quarter_markers:
        ytd_col, ytd_label = ytd_markers[0]
        if ytd_col > 0:
            prior_col_text = " ".join(
                normalize_space(csv_cell_any(rows, r, ytd_col - 1))
                for r in range(start_row, min(len(rows), header_row + 3))
            ).lower()
            if "forms received" in prior_col_text or re.search(r"\breceived\b", prior_col_text):
                ytd_col -= 1
        quarter_label = quarter_markers[0][1]
        return [ytd_label if c >= ytd_col else quarter_label for c in range(max_cols)]

    current = ""
    marker_idx = 0
    for c in range(max_cols):
        while marker_idx < len(markers) and markers[marker_idx][0] <= c:
            current = markers[marker_idx][1]
            marker_idx += 1
        labels[c] = current
    if quarter_markers:
        first_marker_col, first_marker_label = markers[0]
        for c in range(min(first_marker_col, max_cols)):
            labels[c] = first_marker_label
    return labels


def csv_cell_any(rows: list[list[Any]], row: int, col: int) -> Any:
    if row < 0 or row >= len(rows) or col < 0 or col >= len(rows[row]):
        return ""
    return rows[row][col]


def normalize_form_code(value: str) -> str | None:
    text = normalize_space(value)
    match = re.match(r"([A-Z]{1,3}-?\d{2,4}[A-Z]?)(?=\s|,|$)", text, re.I)
    if not match:
        compact = re.sub(r"\s+", "", text)
        match = re.match(r"([A-Z]{1,3}-?\d{2,4}[A-Z]?)", compact, re.I)
    if not match:
        return None
    code = match.group(1).upper()
    if re.match(r"[A-Z]{1,3}-\d{3}\d+$", code):
        code = code[:-1]
    if re.match(r"[A-Z]{1,3}-\d{3}[A-Z]$", code):
        code = code[:-1]
    return code


def clean_decimal(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return str(round(float(value), 3))
    text = normalize_space(value)
    if not text or text in {"-", "—", "–", "N/A"}:
        return None
    text = re.sub(r"[^\d.-]", "", text)
    if not text or text in {"-", ".", "-."}:
        return None
    try:
        return str(round(float(text), 3))
    except ValueError:
        return None


def parse_pdf_preference_country(source: SourceLink, lines: list[str], file_hash: str, snapshot_date: date) -> tuple[list[dict], list[dict]]:
    starts = [
        i for i, line in enumerate(lines)
        if "Current Status for" in line or "Approval Status for" in line
    ]
    if not starts:
        starts = [0]
    starts.append(len(lines))
    facts: list[dict] = []
    countries: dict[str, dict] = {}
    for start, end in zip(starts, starts[1:]):
        block = lines[start:end]
        country_name = "All Countries"
        heading = block[0] if block else ""
        country_match = re.search(r"(?:Current|Approval) Status for\s+(.+)$", heading, re.I)
        if country_match:
            country_name = country_match.group(1).strip()
        country_name = country_name.replace("the Philippines", "Philippines")
        country_code_value, normalized_country = country_from_sheet(country_name)
        countries[country_code_value] = {
            "country_code": country_code_value,
            "country_name": normalized_country,
            "uscis_country_label": normalized_country,
        }
        year_idx = next((i for i, line in enumerate(block[:12]) if len(re.findall(r"\b20\d{2}\b|\b19\d{2}\b", line)) >= 3), None)
        if year_idx is None:
            continue
        years = [int(y) for y in re.findall(r"\b(?:19|20)\d{2}\b", block[year_idx])]
        current_category = "TOTAL"
        approval_section = False
        idx = year_idx + 1
        while idx < len(block):
            line = block[idx]
            lowered = line.lower()
            if lowered.startswith(("table key", "reference", "note(s)", "source(s)", "department of homeland")):
                break
            if "approvals by category" in lowered:
                approval_section = True
                idx += 1
                continue
            parse_line = line
            if status_code(parse_line) == "pending_other" and len(extract_pdf_values(parse_line)) < len(years) and idx + 1 < len(block):
                parse_line = parse_line + " " + block[idx + 1]
                idx += 1
            if "professionals with advanced degrees" in lowered and len(extract_pdf_values(parse_line)) < len(years) and idx + 1 < len(block):
                parse_line = parse_line + " " + block[idx + 1]
                idx += 1

            label = pdf_label(parse_line)
            cat = category_code(label)
            stat = status_code(label)
            if label.lower() == "total petitions":
                stat = "received"
            if cat and not stat and not approval_section:
                current_category = cat
                idx += 1
                continue
            if cat and approval_section:
                current_category = cat
                stat = "approved"
            if stat:
                target_category = cat if cat and stat == "approved" and approval_section else current_category
                values = extract_pdf_values(parse_line)
                for year, token in zip(years, values[:len(years)]):
                    count = clean_count(token)
                    if count is None:
                        continue
                    facts.append(fact_row(
                        file_hash, 0, target_category or "TOTAL", country_code_value, stat,
                        year, None, source, snapshot_date, count, label, str(year),
                        f"L{start + idx + 1}", target_category == "TOTAL", False
                    ))
            idx += 1
    return facts, list(countries.values()) or [{"country_code": "ALL", "country_name": "All Countries", "uscis_country_label": "All Countries"}]


def parse_pdf_all_forms(source: SourceLink, lines: list[str], file_hash: str, snapshot_date: date) -> list[dict]:
    facts: list[dict] = []
    current_category = ""
    pending_form_codes: list[str] = []
    for idx, line in enumerate(lines):
        text = normalize_space(line)
        if not text:
            continue
        if text in {"Family Based", "Employment Based", "Humanitarian", "Citizenship and Naturalization", "Other"}:
            current_category = text
            continue
        form_code = normalize_form_code(text)
        if pending_form_codes and not form_code:
            form_code = pending_form_codes.pop(0)
            text = form_code + " " + text
        if not form_code and not text.lower().startswith("total"):
            continue
        if text.lower().startswith("total"):
            form_code = "TOTAL"
        if form_code not in {"TOTAL", "I-140"}:
            if form_code and len(extract_pdf_values(text)) < 4:
                pending_form_codes.append(form_code)
            continue
        value_text = text
        if form_code and form_code != "TOTAL":
            value_text = re.sub(rf"^\s*{re.escape(form_code)}(?:/\d+[A-Z]?)?\s*", "", value_text, flags=re.I)
        nums = extract_pdf_values(value_text)
        if form_code and len(nums) < 4:
            pending_form_codes.append(form_code)
            continue
        if len(nums) < 4:
            continue
        label = text
        desc = pdf_label(value_text) if form_code != "TOTAL" else "Total - All Forms"

        handled_period = False

        # Older all-forms PDFs list Q1..Qn as received/approved/denied/pending groups.
        group_count = source.report_quarter or 0
        if group_count and len(nums) >= group_count * 4 and (source.report_fiscal_year or 0) <= 2020:
            for q in range(1, group_count + 1):
                base = (q - 1) * 4
                for offset, stat in [(0, "received"), (1, "approved"), (2, "denied"), (3, "pending")]:
                    add_form_fact(facts, file_hash, source, snapshot_date, form_code or "TOTAL", desc, stat,
                                  nums[base + offset], "quarter", q, current_category, label,
                                  f"Q{q} / {stat}", f"L{idx + 1}")
            handled_period = True

        # Newer all-forms PDFs list current quarter plus fiscal-year-to-date.
        if len(nums) >= 11 and not (source.report_fiscal_year and source.report_fiscal_year <= 2020):
            for offset, stat in [(0, "received"), (1, "approved"), (2, "denied"), (4, "pending")]:
                add_form_fact(facts, file_hash, source, snapshot_date, form_code or "TOTAL", desc, stat,
                              nums[offset], "quarter", source.report_quarter, current_category, label,
                              f"Q{source.report_quarter} / {stat}", f"L{idx + 1}")
            for offset, stat in [(6, "received"), (7, "approved"), (8, "denied"), (10, "pending")]:
                add_form_fact(facts, file_hash, source, snapshot_date, form_code or "TOTAL", desc, stat,
                              nums[offset], "ytd", None, current_category, label,
                              f"FYTD / {stat}", f"L{idx + 1}")
            handled_period = True

        # Some FY2021 PDFs expose only the current quarter row: received, approved,
        # denied, completed total, pending, processing time.
        if not handled_period and len(nums) >= 5:
            for offset, stat in [(0, "received"), (1, "approved"), (2, "denied"), (4, "pending")]:
                add_form_fact(facts, file_hash, source, snapshot_date, form_code or "TOTAL", desc, stat,
                              nums[offset], "quarter", source.report_quarter, current_category, label,
                              f"Q{source.report_quarter} / {stat}", f"L{idx + 1}")
    return facts


def add_form_fact(target: list[dict], file_hash: str, source: SourceLink, snapshot_date: date,
                  form_code: str, form_name: str, status: str, value: str, period_scope: str,
                  period_quarter: int | None, form_category: str, row_label: str,
                  column_label: str, address: str) -> None:
    count = clean_decimal(value)
    if count is None:
        return
    target.append({
        "file_hash_sha256": file_hash,
        "sheet_index": 0,
        "form_code": form_code,
        "form_name": form_name,
        "status_code": status,
        "report_fiscal_year": source.report_fiscal_year or "",
        "report_quarter": source.report_quarter or "",
        "snapshot_date": snapshot_date.isoformat(),
        "period_scope": period_scope,
        "period_quarter": period_quarter or "",
        "count_value": count,
        "form_category": form_category,
        "form_description": form_name,
        "raw_row_label": row_label,
        "raw_column_label": column_label,
        "raw_cell_address": address,
        "extraction_method": "script",
        "extraction_confidence": "0.6500",
        "reviewed": "false",
    })


def pdf_label(line: str) -> str:
    cleaned = re.sub(r"\)(\d+)\b", ")", line)
    match = re.search(r"(?<![A-Za-z])(?:\d{1,3}(?:,\d{3})+|\d+|-|D|H)(?![A-Za-z])", cleaned)
    if not match:
        return normalize_space(cleaned)
    return normalize_space(cleaned[:match.start()])


def extract_pdf_values(line: str) -> list[str]:
    cleaned = re.sub(r"\)(\d+)\b", ")", line)
    label = pdf_label(cleaned)
    tail = cleaned[len(label):]
    return re.findall(r"(?<![A-Za-z])(?:\d{1,3}(?:,\d{3})+(?:\.\d+)?|\d+\.\d+|\d+|-|D|H)(?![A-Za-z])", tail)


def parse_fy_quarter_sheet(source: SourceLink, ws, merged: dict[str, tuple[Any, str]], file_hash: str, sheet_index: int, snapshot_date: date) -> list[dict]:
    facts: list[dict] = []
    status_row = find_row_containing(ws, "Received") or 5
    quarter_row = max(1, status_row - 1)
    data_start = status_row + 1
    cohort_fy = source.report_fiscal_year
    if not cohort_fy:
        return facts

    for row in range(data_start, ws.max_row + 1):
        label = normalize_space(cell_value(ws, row, 1, merged))
        if not label:
            continue
        if label.lower().startswith(("table key", "references", "source:")):
            break
        cat = category_code(label)
        if not cat:
            continue
        is_total_row = cat == "TOTAL" or "preference" in label.lower()
        for col in range(2, ws.max_column + 1):
            stat = status_code(str(cell_value(ws, status_row, col, merged)))
            if not stat:
                continue
            count = clean_count(cell_value(ws, row, col, merged))
            if count is None:
                continue
            q_label = normalize_space(cell_value(ws, quarter_row, col, merged))
            q_num = quarter_number(q_label)
            is_total_column = "total" in q_label.lower()
            if q_num and source.report_quarter and q_num > source.report_quarter:
                continue
            facts.append(fact_row(
                file_hash, sheet_index, cat, "ALL", stat, cohort_fy, q_num,
                source, snapshot_date, count, label,
                f"{q_label} / {cell_value(ws, status_row, col, merged)}",
                f"{get_column_letter(col)}{row}", is_total_row, is_total_column
            ))
    return facts


def parse_preference_country_sheet(source: SourceLink, ws, merged: dict[str, tuple[Any, str]], file_hash: str, sheet_index: int, snapshot_date: date, country_code_value: str) -> list[dict]:
    facts: list[dict] = []
    year_row = find_year_header_row(ws, merged)
    if not year_row:
        return facts
    current_category = "TOTAL"
    approval_section = False

    for row in range(year_row + 1, ws.max_row + 1):
        label = normalize_space(cell_value(ws, row, 1, merged))
        if not label:
            continue
        lowered = label.lower()
        if lowered.startswith(("table key", "references", "source:")):
            break
        if "approvals by category" in lowered:
            approval_section = True
            continue
        cat = category_code(label)
        stat = status_code(label)
        if lowered == "total":
            cat = "TOTAL"
            stat = "received"
        if cat and not stat and not approval_section:
            current_category = cat
            continue
        if cat and approval_section:
            current_category = cat
            stat = "approved"
        if not stat:
            continue
        target_category = cat if cat and stat == "approved" and approval_section else current_category
        for col in range(2, ws.max_column + 1):
            header = normalize_space(cell_value(ws, year_row, col, merged))
            if not re.fullmatch(r"\d{4}|TOTAL|Grand Total", header, re.I):
                continue
            count = clean_count(cell_value(ws, row, col, merged))
            if count is None:
                continue
            is_total_column = not header.isdigit()
            if is_total_column:
                continue
            else:
                cohort_fy = int(header)
            if not cohort_fy:
                continue
            facts.append(fact_row(
                file_hash, sheet_index, target_category, country_code_value, stat,
                cohort_fy, None, source, snapshot_date, count, label, header,
                f"{get_column_letter(col)}{row}", target_category == "TOTAL", is_total_column
            ))
    return facts


def find_row_containing(ws, needle: str) -> int | None:
    for row in range(1, min(ws.max_row, 20) + 1):
        values = [normalize_space(ws.cell(row, col).value) for col in range(1, ws.max_column + 1)]
        if any(needle.lower() == value.lower() for value in values):
            return row
    return None


def find_year_header_row(ws, merged: dict[str, tuple[Any, str]]) -> int | None:
    for row in range(1, min(ws.max_row, 20) + 1):
        years = 0
        for col in range(1, ws.max_column + 1):
            value = normalize_space(cell_value(ws, row, col, merged))
            if re.fullmatch(r"\d{4}", value):
                years += 1
        if years >= 3:
            return row
    return None


def fact_row(file_hash: str, sheet_index: int, category: str, country: str, status: str,
             cohort_fy: int, cohort_q: int | None, source: SourceLink, snapshot_date: date,
             count: int, raw_row: str, raw_col: str, address: str,
             is_total_row: bool, is_total_column: bool) -> dict:
    return {
        "file_hash_sha256": file_hash,
        "sheet_index": sheet_index,
        "category_code": category,
        "country_code": country,
        "status_code": status,
        "cohort_fiscal_year": cohort_fy,
        "cohort_quarter": cohort_q or "",
        "report_fiscal_year": source.report_fiscal_year or "",
        "report_quarter": source.report_quarter or "",
        "snapshot_date": snapshot_date.isoformat(),
        "count_value": count,
        "value_type": "count",
        "is_total_row": "true" if is_total_row else "false",
        "is_total_column": "true" if is_total_column else "false",
        "raw_row_label": raw_row,
        "raw_column_label": raw_col,
        "raw_cell_address": address,
        "extraction_method": "script",
        "extraction_confidence": "0.9000",
        "reviewed": "false",
    }


def write_csv(path: Path, rows: list[dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        path.write_text("", encoding="utf-8")
        return
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)


def write_load_sql(staging_dir: Path) -> None:
    def sql_path(name: str) -> str:
        return name.replace("'", "''")

    sql = f"""
BEGIN;

CREATE TEMP TABLE stg_source_files (
    file_hash_sha256 text, source_name text, form_type text, report_family text,
    report_title text, file_name text, source_url text, local_path text,
    report_fiscal_year text, report_quarter text, fiscal_year_label text,
    published_date text, snapshot_date date, parser_version text,
    parse_status text, parse_notes text
);
\\copy stg_source_files FROM '{sql_path("source_files.csv")}' WITH (FORMAT csv, HEADER true, ENCODING 'UTF8')

CREATE TEMP TABLE stg_sheets (
    file_hash_sha256 text, sheet_index int, sheet_name text, max_row int,
    max_column int, detected_table_type text, detection_confidence numeric
);
\\copy stg_sheets FROM '{sql_path("workbook_sheets.csv")}' WITH (FORMAT csv, HEADER true, ENCODING 'UTF8')

CREATE TEMP TABLE stg_cells (
    file_hash_sha256 text, sheet_index int, row_num int, col_num int,
    cell_address text, raw_value text, normalized_value text,
    is_merged boolean, merged_range text
);
\\copy stg_cells FROM '{sql_path("sheet_cells.csv")}' WITH (FORMAT csv, HEADER true, ENCODING 'UTF8')

CREATE TEMP TABLE stg_countries (
    country_code text, country_name text, uscis_country_label text
);
\\copy stg_countries FROM '{sql_path("countries.csv")}' WITH (FORMAT csv, HEADER true, ENCODING 'UTF8')

CREATE TEMP TABLE stg_facts (
    file_hash_sha256 text, sheet_index int, category_code text, country_code text,
    status_code text, cohort_fiscal_year int, cohort_quarter text,
    report_fiscal_year text, report_quarter text, snapshot_date date,
    count_value int, value_type text, is_total_row boolean, is_total_column boolean,
    raw_row_label text, raw_column_label text, raw_cell_address text,
    extraction_method text, extraction_confidence numeric, reviewed boolean
);
\\copy stg_facts FROM '{sql_path("facts.csv")}' WITH (FORMAT csv, HEADER true, ENCODING 'UTF8')

CREATE TEMP TABLE stg_form_facts (
    file_hash_sha256 text, sheet_index int, form_code text, form_name text,
    status_code text, report_fiscal_year text, report_quarter text,
    snapshot_date date, period_scope text, period_quarter text,
    count_value numeric, form_category text, form_description text,
    raw_row_label text, raw_column_label text, raw_cell_address text,
    extraction_method text, extraction_confidence numeric, reviewed boolean
);
\\copy stg_form_facts FROM '{sql_path("form_facts.csv")}' WITH (FORMAT csv, HEADER true, ENCODING 'UTF8')

INSERT INTO uscis_dim.countries (country_code, country_name, uscis_country_label)
SELECT DISTINCT country_code, country_name, uscis_country_label
FROM stg_countries
ON CONFLICT (country_code) DO UPDATE
SET country_name = EXCLUDED.country_name,
    uscis_country_label = EXCLUDED.uscis_country_label;

INSERT INTO uscis_raw.source_files (
    file_hash_sha256, source_name, form_type, report_family, report_title,
    file_name, source_url, local_path, report_fiscal_year, report_quarter,
    fiscal_year_label, published_date, snapshot_date, parser_version,
    parse_status, parse_notes
)
SELECT
    file_hash_sha256, source_name, form_type, report_family, report_title,
    file_name, source_url, local_path,
    NULLIF(report_fiscal_year, '')::int,
    NULLIF(report_quarter, '')::int,
    fiscal_year_label,
    NULLIF(published_date, '')::date,
    snapshot_date, parser_version, parse_status, parse_notes
FROM stg_source_files
ON CONFLICT (file_hash_sha256) DO UPDATE
SET local_path = EXCLUDED.local_path,
    snapshot_date = EXCLUDED.snapshot_date,
    parser_version = EXCLUDED.parser_version,
    parse_status = EXCLUDED.parse_status,
    parse_notes = EXCLUDED.parse_notes;

INSERT INTO uscis_raw.workbook_sheets (
    source_file_id, sheet_name, sheet_index, max_row, max_column,
    detected_table_type, detection_confidence
)
SELECT sf.id, s.sheet_name, s.sheet_index, s.max_row, s.max_column,
       s.detected_table_type, s.detection_confidence
FROM stg_sheets s
JOIN uscis_raw.source_files sf ON sf.file_hash_sha256 = s.file_hash_sha256
ON CONFLICT (source_file_id, sheet_index) DO UPDATE
SET sheet_name = EXCLUDED.sheet_name,
    max_row = EXCLUDED.max_row,
    max_column = EXCLUDED.max_column,
    detected_table_type = EXCLUDED.detected_table_type,
    detection_confidence = EXCLUDED.detection_confidence;

INSERT INTO uscis_raw.sheet_cells (
    sheet_id, row_num, col_num, cell_address, raw_value,
    normalized_value, is_merged, merged_range
)
SELECT ws.id, c.row_num, c.col_num, c.cell_address, c.raw_value,
       c.normalized_value, c.is_merged, c.merged_range
FROM stg_cells c
JOIN uscis_raw.source_files sf ON sf.file_hash_sha256 = c.file_hash_sha256
JOIN uscis_raw.workbook_sheets ws ON ws.source_file_id = sf.id AND ws.sheet_index = c.sheet_index
ON CONFLICT (sheet_id, row_num, col_num) DO UPDATE
SET raw_value = EXCLUDED.raw_value,
    normalized_value = EXCLUDED.normalized_value,
    is_merged = EXCLUDED.is_merged,
    merged_range = EXCLUDED.merged_range;

DELETE FROM uscis_fact.i140_status_counts f
USING uscis_raw.source_files sf, stg_source_files stg
WHERE f.source_file_id = sf.id
  AND sf.file_hash_sha256 = stg.file_hash_sha256;

DELETE FROM uscis_fact.form_status_counts f
USING uscis_raw.source_files sf, stg_source_files stg
WHERE f.source_file_id = sf.id
  AND sf.file_hash_sha256 = stg.file_hash_sha256;

INSERT INTO uscis_fact.i140_status_counts (
    source_file_id, sheet_id, form_id, report_family_id, category_id,
    country_id, status_id, cohort_fiscal_year, cohort_quarter,
    report_fiscal_year, report_quarter, snapshot_date, count_value,
    value_type, is_total_row, is_total_column, raw_row_label,
    raw_column_label, raw_cell_address, extraction_method,
    extraction_confidence, reviewed
)
SELECT
    sf.id, ws.id, frm.id, rf.id, pc.id, co.id, cs.id,
    f.cohort_fiscal_year, NULLIF(f.cohort_quarter, '')::int,
    NULLIF(f.report_fiscal_year, '')::int,
    NULLIF(f.report_quarter, '')::int,
    f.snapshot_date, f.count_value, f.value_type,
    f.is_total_row, f.is_total_column, f.raw_row_label,
    f.raw_column_label, f.raw_cell_address, f.extraction_method,
    f.extraction_confidence, f.reviewed
FROM stg_facts f
JOIN uscis_raw.source_files sf ON sf.file_hash_sha256 = f.file_hash_sha256
JOIN uscis_raw.workbook_sheets ws ON ws.source_file_id = sf.id AND ws.sheet_index = f.sheet_index
JOIN uscis_dim.forms frm ON frm.form_code = 'I-140'
JOIN uscis_dim.report_families rf ON rf.report_family_code = sf.report_family
JOIN uscis_dim.preference_categories pc ON pc.category_code = f.category_code
JOIN uscis_dim.countries co ON co.country_code = f.country_code
JOIN uscis_dim.case_statuses cs ON cs.status_code = f.status_code
ON CONFLICT DO NOTHING;

INSERT INTO uscis_dim.forms (form_code, form_name)
SELECT
    form_code,
    MAX(COALESCE(NULLIF(form_name, ''), NULLIF(form_description, ''), form_code)) AS form_name
FROM stg_form_facts
WHERE form_code IS NOT NULL AND form_code <> ''
GROUP BY form_code
ON CONFLICT (form_code) DO UPDATE
SET form_name = COALESCE(EXCLUDED.form_name, uscis_dim.forms.form_name);

INSERT INTO uscis_fact.form_status_counts (
    source_file_id, sheet_id, form_id, status_id, report_fiscal_year,
    report_quarter, snapshot_date, period_scope, period_quarter,
    count_value, form_category, form_description, raw_row_label,
    raw_column_label, raw_cell_address, extraction_method,
    extraction_confidence, reviewed
)
SELECT
    sf.id, ws.id, frm.id, cs.id,
    NULLIF(f.report_fiscal_year, '')::int,
    NULLIF(f.report_quarter, '')::int,
    f.snapshot_date, f.period_scope,
    NULLIF(f.period_quarter, '')::int,
    f.count_value, f.form_category, f.form_description,
    f.raw_row_label, f.raw_column_label, f.raw_cell_address,
    f.extraction_method, f.extraction_confidence, f.reviewed
FROM stg_form_facts f
JOIN uscis_raw.source_files sf ON sf.file_hash_sha256 = f.file_hash_sha256
JOIN uscis_raw.workbook_sheets ws ON ws.source_file_id = sf.id AND ws.sheet_index = f.sheet_index
JOIN uscis_dim.forms frm ON frm.form_code = f.form_code
JOIN uscis_dim.case_statuses cs ON cs.status_code = f.status_code
ON CONFLICT DO NOTHING;

COMMIT;
"""
    (staging_dir / "latest_load.sql").write_text(sql.strip() + "\n", encoding="utf-8")


def run(root: Path) -> None:
    env = load_env(root)
    index_url = env.get("USCIS_DATA_URLS") or env.get("USCIS_DATA_URL") or "\n".join(
        [
            "https://www.uscis.gov/tools/reports-and-studies/immigration-and-citizenship-data?topic_id%5B%5D=33628&ddt_mon=&ddt_yr=&query=I-140&items_per_page=100",
            "https://www.uscis.gov/tools/reports-and-studies/immigration-and-citizenship-data?topic_id%5B%5D=33667&ddt_mon=&ddt_yr=&query=I-140&items_per_page=100",
        ]
    )
    explicit_snapshot = date.fromisoformat(env["SNAPSHOT_DATE"]) if env.get("SNAPSHOT_DATE") else None
    probe_start_year = int(env.get("PROBE_START_YEAR") or "2014")
    probe_end_year = int(env.get("PROBE_END_YEAR") or str(date.today().year))
    raw_dir = root / "data" / "raw"
    staging_dir = root / "data" / "staging"

    allow_network = (env.get("ALLOW_NETWORK_DISCOVERY") or "0").lower() in {"1", "true", "yes"}
    sources = []
    if allow_network:
        sources.extend(discover_sources_with_probes(index_url, probe_start_year, probe_end_year))
    sources.extend(discover_local_hand_sources(root))
    if not sources:
        raise SystemExit("No supported USCIS I-140 XLSX links found.")

    all_source_rows: list[dict] = []
    all_sheets: list[dict] = []
    all_cells: list[dict] = []
    all_facts: list[dict] = []
    all_form_facts: list[dict] = []
    all_countries: list[dict] = [{"country_code": "ALL", "country_name": "All Countries", "uscis_country_label": "All Countries"}]
    seen_hashes: set[str] = set()

    for source in sources:
        try:
            path, file_hash = download_file(source, raw_dir)
        except Exception as exc:
            print(f"{source.file_name}: download failed ({exc})")
            continue
        if file_hash in seen_hashes:
            print(f"{source.file_name}: duplicate file hash, skipped")
            continue
        seen_hashes.add(file_hash)
        snapshot = explicit_snapshot or fiscal_quarter_end(source.report_fiscal_year, source.report_quarter)
        source_rows, sheets, cells, facts, countries, form_facts = parse_source_file(source, path, file_hash, snapshot)
        all_source_rows.extend(source_rows)
        all_sheets.extend(sheets)
        all_cells.extend(cells)
        all_facts.extend(facts)
        all_form_facts.extend(form_facts)
        all_countries.extend(countries)
        print(f"{source.file_name}: {len(facts)} i140 facts, {len(form_facts)} form facts")

    country_seen: dict[str, dict] = {}
    for row in all_countries:
        country_seen[row["country_code"]] = row

    write_csv(staging_dir / "source_files.csv", all_source_rows)
    write_csv(staging_dir / "workbook_sheets.csv", all_sheets)
    write_csv(staging_dir / "sheet_cells.csv", all_cells)
    write_csv(staging_dir / "countries.csv", list(country_seen.values()))
    write_csv(staging_dir / "facts.csv", all_facts)
    write_csv(staging_dir / "form_facts.csv", all_form_facts)
    write_load_sql(staging_dir)

    print(f"Discovered files: {len(sources)}")
    print(f"Facts staged: {len(all_facts)}")
    print(f"Form facts staged: {len(all_form_facts)}")
    print(f"Load SQL: {staging_dir / 'latest_load.sql'}")


def main() -> None:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    parser = argparse.ArgumentParser()
    parser.add_argument("--root", type=Path, default=Path.cwd())
    args = parser.parse_args()
    run(args.root.resolve())


if __name__ == "__main__":
    main()
